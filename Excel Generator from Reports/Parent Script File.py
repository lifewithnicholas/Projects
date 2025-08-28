import pandas as pd
import re
from openpyxl import load_workbook

def clean_text(text: str) -> str:
    """Remove unwanted characters, extra spaces, and standardize casing."""
    if pd.isna(text):
        return ""
    text = str(text)
    text = re.sub(r'\s+', ' ', text)  # remove multiple spaces/newlines
    text = text.strip()
    return text.title()  # standardize casing (optional)

def process_report(input_file: str, output_file: str, sheet_name: str = "CleanedData"):
    """
    Read raw report data, clean it, and export to Excel.
    
    Parameters:
        input_file (str): Path to CSV or Excel report file.
        output_file (str): Path to save cleaned Excel file.
        sheet_name (str): Name of the sheet to save cleaned data.
    """

    # Step 1: Load raw data
    if input_file.endswith(".csv"):
        df = pd.read_csv(input_file)
    elif input_file.endswith((".xls", ".xlsx")):
        df = pd.read_excel(input_file)
    else:
        raise ValueError("Unsupported file format. Use CSV or Excel.")

    # Step 2: Clean column names
    df.columns = [clean_text(col) for col in df.columns]

    # Step 3: Clean text fields in the DataFrame
    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].apply(clean_text)

    # Step 4: Remove duplicates and blank rows
    df = df.drop_duplicates().dropna(how="all")

    # Step 5: Example of creating meaningful derived columns
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Year"] = df["Date"].dt.year
        df["Month"] = df["Date"].dt.month_name()

    # Step 6: Export to Excel (append if file exists, else create new)
    try:
        book = load_workbook(output_file)
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
            writer.book = book
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]  # overwrite old sheet
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    except FileNotFoundError:
        df.to_excel(output_file, index=False, sheet_name=sheet_name)

    print(f"✅ Cleaned data saved to {output_file} in sheet '{sheet_name}'")

# Example usage
if __name__ == "__main__":
    process_report("raw_report.csv", "cleaned_report.xlsx")
